#!/usr/bin/env python3

import argparse
import csv
import json
import sys
from pathlib import Path
from zipfile import BadZipFile, ZipFile, is_zipfile


HEADER_INPUT_SHAPES = "inputshapes"
HEADER_TYPE_CANDIDATES = ("type", "optype")
DEFAULT_NAME_PATTERNS = ("kernel_details", "op_analysis_details")
CSV_ENCODINGS = ("utf-8-sig", "utf-8", "gb18030")


def normalize_header(value):
    text = "" if value is None else str(value)
    return "".join(ch for ch in text.strip().lower() if ch not in " _-")


def normalize_type(value):
    text = "" if value is None else str(value)
    return "".join(text.strip().lower().split())


def normalize_shapes(value):
    text = "" if value is None else str(value)
    return (
        text.strip()
        .replace("\"", "")
        .replace("'", "")
        .replace("，", ",")
        .replace("；", ";")
        .replace("：", ":")
        .replace("[", "")
        .replace("]", "")
        .replace("(", "")
        .replace(")", "")
    )


def parse_dims(part):
    dims = []
    for token in part.split(","):
        token = token.strip()
        if not token:
            continue
        dims.append(int(token))
    return dims


def parse_shapes(shapes):
    normalized = normalize_shapes(shapes)
    pieces = [piece.strip() for piece in normalized.split(";") if piece.strip()]
    
    # 允许包含 3 个或更多部分（如包含 Bias 的情况），仅取前两部分用于解析 M, N, K
    pieces = pieces[:2]
    
    if len(pieces) != 2:
        raise ValueError("expected at least two shape groups separated by ';'")

    left = parse_dims(pieces[0])
    right = parse_dims(pieces[1])

    if len(left) == 2 and len(right) == 2:
        m = left[0]
        k = left[1]
        n = right[0]
        if right[1] != k:
            raise ValueError(f"k mismatch between groups: {k} != {right[1]}")
        return {"m": m, "n": n, "k": k, "rule": "basic-2x2"}

    if len(left) == 2 and len(right) == 4:
        m = left[0]
        k = left[1]
        derived_k = right[1] * right[2]
        if derived_k != k:
            raise ValueError(f"packed k mismatch: {k} != {derived_k}")
        n = right[0] * right[3]
        return {"m": m, "n": n, "k": k, "rule": "packed-2x4"}

    if len(left) == 1 and len(right) == 4:
        m = left[0]
        k = right[1] * right[2]
        n = right[0] * right[3]
        return {"m": m, "n": n, "k": k, "rule": "legacy-1x4"}

    raise ValueError(
        f"unsupported shape layout: left has {len(left)} dims, right has {len(right)} dims"
    )


def require_openpyxl():
    try:
        from openpyxl import Workbook, load_workbook
    except ModuleNotFoundError as exc:
        raise RuntimeError(
            "openpyxl is required for Excel workbook input or output. Install it with: "
            "python3 -m pip install openpyxl"
        ) from exc
    return Workbook, load_workbook


def matches_filename(path, patterns):
    filename = path.name.lower()
    return any(pattern in filename for pattern in patterns)


def iter_candidate_files(input_path, patterns):
    if input_path.is_file():
        if not matches_filename(input_path, patterns):
            raise ValueError(
                "input file name must contain kernel_details or op_analysis_details"
            )
        return [input_path]

    if not input_path.is_dir():
        raise ValueError(f"input path not found: {input_path}")

    files = [
        path
        for path in input_path.rglob("*")
        if path.is_file()
        and matches_filename(path, patterns)
    ]
    if not files:
        raise ValueError("no matching kernel_details or op_analysis_details files were found")
    return sorted(files)


def find_required_columns(headers):
    header_map = {}
    for index, value in enumerate(headers):
        normalized = normalize_header(value)
        if normalized and normalized not in header_map:
            header_map[normalized] = index

    type_key = next((key for key in HEADER_TYPE_CANDIDATES if key in header_map), None)
    missing = []
    if type_key is None:
        missing.append("Type or Op Type")
    if HEADER_INPUT_SHAPES not in header_map:
        missing.append("Input Shapes")
    if missing:
        raise ValueError(f"missing required header(s) in row 1: {', '.join(missing)}")

    return header_map[type_key], header_map[HEADER_INPUT_SHAPES]


def is_target_type(value):
    return "matmul" in normalize_type(value)


def get_cell(row, index):
    if index >= len(row):
        return None
    return row[index]


def build_entry(source_path, file_name, sheet_name, row_num, type_column, type_value, shapes_value):
    return {
        "source_path": str(source_path),
        "file_name": file_name,
        "sheet_name": sheet_name,
        "row_num": row_num,
        "type_column": type_column,
        "type_value": "" if type_value is None else str(type_value),
        "input_shapes": "" if shapes_value is None else str(shapes_value),
        "m": None,
        "n": None,
        "k": None,
        "rule": None,
        "error": None,
    }


def extract_from_table(headers, rows, source_path, file_name, sheet_name):
    type_index, shapes_index = find_required_columns(headers)
    type_column = "Op Type" if normalize_header(headers[type_index]) == "optype" else "Type"
    results = []

    for row_num, row in enumerate(rows, start=2):
        type_value = get_cell(row, type_index)
        if not is_target_type(type_value):
            continue

        shapes_value = get_cell(row, shapes_index)
        entry = build_entry(
            source_path=source_path,
            file_name=file_name,
            sheet_name=sheet_name,
            row_num=row_num,
            type_column=type_column,
            type_value=type_value,
            shapes_value=shapes_value,
        )
        try:
            entry.update(parse_shapes(entry["input_shapes"]))
        except Exception as exc:
            entry["error"] = str(exc)
        results.append(entry)

    return results


def parse_text_rows(text):
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t|")
    except csv.Error:
        dialect = csv.excel
    return list(csv.reader(text.splitlines(), dialect))


def load_text_rows(path):
    last_error = None
    for encoding in CSV_ENCODINGS:
        try:
            text = path.read_text(encoding=encoding)
            # handle triple-quoted fields which are common in some exports
            if '"""' in text:
                text = text.replace('"""', '"')
            return parse_text_rows(text)
        except UnicodeDecodeError as exc:
            last_error = exc
    raise RuntimeError(f"failed to decode text table file {path}: {last_error}")


def extract_from_text_table(path):
    rows = load_text_rows(path)
    if not rows:
        return []
    return extract_from_table(
        headers=rows[0],
        rows=rows[1:],
        source_path=path.resolve(),
        file_name=path.name,
        sheet_name="",
    )


def extract_from_excel(path):
    _, load_workbook = require_openpyxl()
    with path.open("rb") as handle:
        workbook = load_workbook(handle, data_only=True)
        try:
            results = []
            for sheet in workbook.worksheets:
                header_row = next(
                    sheet.iter_rows(min_row=1, max_row=1, values_only=True),
                    (),
                )
                data_rows = list(sheet.iter_rows(min_row=2, values_only=True))
                results.extend(
                    extract_from_table(
                        headers=list(header_row),
                        rows=data_rows,
                        source_path=path.resolve(),
                        file_name=path.name,
                        sheet_name=sheet.title,
                    )
                )
            return results
        finally:
            workbook.close()


def is_excel_workbook(path):
    if not is_zipfile(path):
        return False

    try:
        with ZipFile(path) as archive:
            names = set(archive.namelist())
    except (OSError, BadZipFile):
        return False

    return "[Content_Types].xml" in names and "xl/workbook.xml" in names


def extract_rows(input_path, patterns):
    results = []
    for path in iter_candidate_files(input_path, patterns):
        if is_excel_workbook(path):
            results.extend(extract_from_excel(path))
        else:
            results.extend(extract_from_text_table(path))
    return results


def write_csv(path, rows):
    fieldnames = [
        "source_path",
        "file_name",
        "sheet_name",
        "row_num",
        "type_column",
        "type_value",
        "input_shapes",
        "m",
        "n",
        "k",
        "rule",
        "error",
    ]
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def write_excel(path, rows):
    Workbook, _ = require_openpyxl()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "matmul_mnk"
    fieldnames = [
        "source_path",
        "file_name",
        "sheet_name",
        "row_num",
        "type_column",
        "type_value",
        "input_shapes",
        "m",
        "n",
        "k",
        "rule",
        "error",
    ]
    sheet.append(fieldnames)
    for row in rows:
        sheet.append([row.get(field) for field in fieldnames])
    workbook.save(path)


def write_output(path, rows):
    suffix = path.suffix.lower()
    if suffix == ".json":
        path.write_text(json.dumps(rows, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
        return
    if suffix == ".csv":
        write_csv(path, rows)
        return
    if suffix in {".xlsx", ".xlsm"}:
        write_excel(path, rows)
        return
    raise ValueError("output path must end with .json, .csv, .xlsx, or .xlsm")


def build_parser():
    parser = argparse.ArgumentParser(
        description=(
            "Given a file path or directory path, find files whose names contain "
            "kernel_details or op_analysis_details and extract M, N, and K from "
            "rows whose op type contains matmul."
        )
    )
    parser.add_argument(
        "--input",
        required=True,
        help=(
            "A file path or a directory path. If a directory is provided, scan it "
            "recursively for files whose names contain kernel_details or "
            "op_analysis_details."
        ),
    )
    parser.add_argument(
        "--pattern",
        action="append",
        default=[],
        help="Optional extra case-insensitive filename substring to match. Repeat as needed.",
    )
    parser.add_argument(
        "--output",
        help="Optional output path ending in .json, .csv, .xlsx, or .xlsm. Default is JSON to stdout.",
    )
    return parser


def main():
    args = build_parser().parse_args()
    input_path = Path(args.input).expanduser().resolve()
    patterns = tuple(pattern.lower() for pattern in (list(DEFAULT_NAME_PATTERNS) + args.pattern))

    try:
        rows = extract_rows(input_path, patterns)
        if args.output:
            output_path = Path(args.output).expanduser().resolve()
            write_output(output_path, rows)
        else:
            print(json.dumps(rows, indent=2, ensure_ascii=False))
    except Exception as exc:
        print(str(exc), file=sys.stderr)
        return 1

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
